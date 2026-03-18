"""Workbook inspector: reads sheet headers and generates mapping templates."""

import json
from pathlib import Path

import openpyxl

from spark_pdm_generator.parsers.column_mapper import generate_mapping_template


def inspect_workbook(workbook_path: Path) -> dict[str, list[str]]:
    """Read all sheet names and their column headers from a workbook.

    Returns:
        Dict mapping sheet tab names to lists of column header strings.
    """
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_headers: dict[str, list[str]] = {}

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if first_row:
                headers = [str(h).strip() if h else "" for h in first_row]
                sheet_headers[sheet_name] = [h for h in headers if h]
            else:
                sheet_headers[sheet_name] = []
    finally:
        wb.close()

    return sheet_headers


def generate_and_save_mapping(
    workbook_path: Path, output_path: Path
) -> dict[str, list[str]]:
    """Inspect a workbook and generate a column mapping template JSON file.

    Args:
        workbook_path: Path to the input Excel workbook.
        output_path: Path where the mapping JSON will be written.

    Returns:
        The sheet_headers dict for display purposes.
    """
    sheet_headers = inspect_workbook(workbook_path)
    template = generate_mapping_template(sheet_headers)

    with open(output_path, "w") as f:
        json.dump(template, f, indent=2)

    return sheet_headers
