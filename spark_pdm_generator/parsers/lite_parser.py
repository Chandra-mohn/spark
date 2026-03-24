"""Lite parser: reads a simplified 3-sheet workbook (Entities, Attributes, Relationships).

This parser handles input workbooks with minimal metadata -- entity names,
attributes with datatypes/sizes, primary keys, and relationships. It maps
the non-standard sheet/column names to the standard LogicalModel.

Designed to be disposable -- replace with the full ExcelParser when
comprehensive input data becomes available.
"""

import re
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from spark_pdm_generator.models.logical import (
    Attribute,
    Cardinality,
    Config,
    Entity,
    EntityType,
    LogicalModel,
    Relationship,
)
from spark_pdm_generator.parsers.utils import (
    build_config_from_dict as _build_config_from_dict,
    parse_enum as _parse_enum,
    parse_int as _parse_int,
)


class LiteParseError(Exception):
    """Raised when the lite input workbook has structural issues."""

    pass


# Datatype mapping: input Datatype -> (normalized_logical_type, default_max_length_bytes)
DATATYPE_MAP: dict[str, tuple[str, int]] = {
    "CHARACTERS": ("VARCHAR", 0),       # length from Length column
    "TEXT": ("VARCHAR", 0),             # length from Length column
    "NUMBER": ("DECIMAL", 16),          # precision from Length column
    "DECIMAL": ("DECIMAL", 16),         # precision from Length column
    "INTEGER": ("INTEGER", 4),
    "LONG INTEGER": ("BIGINT", 8),
    "DATE": ("DATE", 4),
    "TIMESTAMP": ("TIMESTAMP", 8),
    "BOOLEAN": ("BOOLEAN", 1),
    "BINARY": ("BINARY", 0),           # length from Length column
}

# Storage size in bytes for each normalized logical type
STORAGE_BYTES: dict[str, int] = {
    "VARCHAR": 0,       # uses max_length
    "DECIMAL": 16,      # Parquet fixed-length byte array
    "INTEGER": 4,
    "BIGINT": 8,
    "DATE": 4,
    "TIMESTAMP": 8,
    "BOOLEAN": 1,
    "BINARY": 0,        # uses max_length
}


class LiteParser:
    """Reads a simplified 3-sheet Excel workbook and produces a LogicalModel."""

    def __init__(self, flip_composition: bool = False) -> None:
        self.errors: list[str] = []
        self.warnings: list[str] = []
        self.flip_composition = flip_composition

    def parse(self, workbook_path: Path) -> LogicalModel:
        """Parse the lite input workbook into a LogicalModel.

        Expected sheets: Entities, Attributes, Relationships

        Args:
            workbook_path: Path to the Excel file.

        Returns:
            A populated LogicalModel.

        Raises:
            LiteParseError: If required sheets or fields are missing.
        """
        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

        try:
            entities = self._parse_entities(wb)
            # Build canonical name map for cross-sheet normalization
            self._entity_name_map = _build_name_map(
                [e.entity_name for e in entities]
            )
            attributes = self._parse_attributes(wb)
            relationships = self._parse_relationships(wb, attributes)
            config = self._parse_config(wb)
        finally:
            wb.close()

        # Warn on duplicate entity names
        seen_entities: set[str] = set()
        for e in entities:
            if e.entity_name in seen_entities:
                self.warnings.append(
                    f"Duplicate entity name: '{e.entity_name}'"
                )
            seen_entities.add(e.entity_name)

        # Warn on duplicate (entity, attribute) pairs
        seen_attrs: set[tuple[str, str]] = set()
        for a in attributes:
            key = (a.entity_name, a.attribute_name)
            if key in seen_attrs:
                self.warnings.append(
                    f"Duplicate attribute: '{a.entity_name}.{a.attribute_name}'"
                )
            seen_attrs.add(key)

        # Validate: warn about entities with zero attributes
        entity_names = {e.entity_name for e in entities}
        attr_entities = {a.entity_name for a in attributes}
        for name in sorted(entity_names - attr_entities):
            self.warnings.append(
                f"Entity '{name}' has 0 attributes -- check EntityName "
                f"spelling between Entities and Attributes sheets"
            )

        if self.errors:
            raise LiteParseError(
                "Parsing errors:\n" + "\n".join(f"  - {e}" for e in self.errors)
            )

        return LogicalModel(
            entities=entities,
            attributes=attributes,
            relationships=relationships,
            config=config,
        )

    # ------------------------------------------------------------------
    # Entity parsing
    # ------------------------------------------------------------------

    def _parse_entities(self, wb: openpyxl.Workbook) -> list[Entity]:
        """Parse the Entities sheet.

        Entity names are stripped of leading/trailing whitespace.
        """
        ws = self._get_sheet(wb, "Entities", required=True)
        if not ws:
            return []

        rows = self._read_sheet_rows(ws)
        entities = []

        for i, row in enumerate(rows, start=2):
            entity_name = _str(row.get("EntityName")).strip()
            if not entity_name:
                continue

            row_count = _parse_int(row.get("RowCount"))

            # Parse optional EntityType override
            raw_type = _str(row.get("EntityType")).strip().upper()
            entity_type = EntityType.UNKNOWN
            if raw_type:
                try:
                    entity_type = EntityType(raw_type)
                except ValueError:
                    self.warnings.append(
                        f"Entities row {i}: unknown EntityType '{raw_type}' "
                        f"for '{entity_name}', will auto-classify"
                    )

            entity = Entity(
                entity_name=entity_name,
                entity_type=entity_type,
                description=_str(row.get("Comment", "")),
                estimated_row_count=row_count,
                domain=_str(row.get("CollectionName", "general")) or "general",
            )
            entities.append(entity)

        if not entities:
            self.errors.append("Entities sheet is empty or has no valid rows.")

        return entities

    # ------------------------------------------------------------------
    # Attribute parsing
    # ------------------------------------------------------------------

    def _parse_attributes(self, wb: openpyxl.Workbook) -> list[Attribute]:
        """Parse the Attributes sheet."""
        ws = self._get_sheet(wb, "Attributes", required=True)
        if not ws:
            return []

        rows = self._read_sheet_rows(ws)
        attributes = []

        for i, row in enumerate(rows, start=2):
            raw_entity_name = _str(row.get("EntityName")).strip()
            attr_code = _str(row.get("Code"))
            raw_datatype = _str(row.get("Datatype", "")).strip().upper()
            length = _parse_int(row.get("Length"))

            if not raw_entity_name or not attr_code:
                continue

            # Normalize entity name to match Entities sheet
            entity_name = self._entity_name_map.get(
                raw_entity_name.lower(), raw_entity_name
            )
            if entity_name != raw_entity_name:
                self.warnings.append(
                    f"Attributes row {i}: normalized EntityName "
                    f"'{raw_entity_name}' -> '{entity_name}'"
                )

            if not raw_datatype:
                self.warnings.append(
                    f"Attributes row {i}: missing Datatype for "
                    f"'{entity_name}.{attr_code}', defaulting to VARCHAR"
                )
                raw_datatype = "CHARACTERS"

            # Map datatype
            type_info = DATATYPE_MAP.get(raw_datatype)
            if not type_info:
                self.warnings.append(
                    f"Attributes row {i}: unknown Datatype '{raw_datatype}' "
                    f"for '{entity_name}.{attr_code}', defaulting to VARCHAR"
                )
                type_info = ("VARCHAR", 0)

            logical_type, default_bytes = type_info

            # Determine precision/scale for numeric types
            precision: Optional[int] = None
            scale: Optional[int] = None
            max_length: Optional[int] = None

            if logical_type == "DECIMAL":
                precision = length if length else 18
                # Number -> scale=0 (whole number), Decimal -> scale=2 (financial)
                if raw_datatype == "NUMBER":
                    scale = 0
                else:
                    scale = 2
            elif logical_type in ("VARCHAR", "BINARY"):
                max_length = length if length else 256
            # INTEGER, BIGINT, DATE, TIMESTAMP, BOOLEAN -- no length needed

            # Parse flags
            is_pk = _str(row.get("PrimaryIdentifierFlag", "")).strip().upper() == "Y"
            is_fk = _str(row.get("ForeignIdentifierFlag", "")).strip().upper() == "Y"
            nullable = _str(row.get("MandatoryFlag", "")).strip().upper() != "Y"
            raw_fk_parent = _str(row.get("ForeignIdentifierParentEntityName")).strip()
            fk_parent = self._entity_name_map.get(
                raw_fk_parent.lower(), raw_fk_parent
            ) if raw_fk_parent else ""

            attribute = Attribute(
                entity_name=entity_name,
                attribute_name=attr_code,
                logical_data_type=logical_type,
                precision=precision,
                scale=scale,
                max_length=max_length,
                nullable=nullable,
                is_primary_key=is_pk,
                is_foreign_key=is_fk,
                fk_references=fk_parent if fk_parent and fk_parent != "N/A" else None,
                description=_str(row.get("Comment", "")),
            )
            attributes.append(attribute)

        return attributes

    # ------------------------------------------------------------------
    # Relationship parsing
    # ------------------------------------------------------------------

    def _parse_relationships(
        self, wb: openpyxl.Workbook, attributes: list[Attribute]
    ) -> list[Relationship]:
        """Parse the Relationships sheet.

        JoinAttributes format: "ParentEntity.Attr Name = ChildEntity.Attr Name"
        Uses display names (EntityName, Attribute Name) which must be mapped
        back to Code via the attributes list.
        """
        ws = self._get_sheet(wb, "Relationships", required=True)
        if not ws:
            return []

        # Build lookup: (EntityName, Attribute Name) -> Attribute Code
        # The "Name" column in attributes has the display name
        rows = self._read_sheet_rows(ws)
        relationships = []

        for i, row in enumerate(rows, start=2):
            raw_parent = _str(row.get("ParentEntity")).strip()
            raw_child = _str(row.get("ChildEntity")).strip()
            rel_type = _str(row.get("RelationshipType", "1:n")).strip()
            join_attrs = _str(row.get("JoinAttributes", ""))
            stereotype = _str(row.get("Stereotype", "")).strip().lower()

            if not raw_parent or not raw_child:
                continue

            # Normalize entity names to match Entities sheet
            parent_entity = self._entity_name_map.get(
                raw_parent.lower(), raw_parent
            )
            child_entity = self._entity_name_map.get(
                raw_child.lower(), raw_child
            )

            # Parse cardinality
            cardinality, cardinality_defaulted = _parse_cardinality(rel_type)
            if cardinality_defaulted:
                self.warnings.append(
                    f"Relationships row {i}: unrecognized RelationshipType "
                    f"'{rel_type}' for {parent_entity} -> {child_entity}, "
                    f"defaulting to 1:N"
                )

            # Parse join attributes: "Entity.AttrName = Entity.AttrName"
            parent_cols, child_cols = _parse_join_attributes(join_attrs)

            if not parent_cols or not child_cols:
                self.warnings.append(
                    f"Relationships row {i}: could not parse JoinAttributes "
                    f"'{join_attrs}' for {parent_entity} -> {child_entity}"
                )

            is_composition = stereotype == "composition"

            # Flip composition direction if requested -- when the input data
            # has subdocument as ParentEntity and collection as ChildEntity
            if is_composition and self.flip_composition:
                parent_entity, child_entity = child_entity, parent_entity
                parent_cols, child_cols = child_cols, parent_cols
                self.warnings.append(
                    f"Relationships row {i}: flipped composition "
                    f"direction -- parent is now '{parent_entity}', "
                    f"child is now '{child_entity}'"
                )

            description = ""
            if is_composition:
                description = "composition: embedded array in source MongoDB document"

            rel = Relationship(
                parent_entity=parent_entity,
                child_entity=child_entity,
                cardinality=cardinality,
                parent_key_columns=parent_cols,
                child_key_columns=child_cols,
                is_identifying=is_composition,
                description=description,
            )
            relationships.append(rel)

        return relationships

    # ------------------------------------------------------------------
    # Config parsing
    # ------------------------------------------------------------------

    def _parse_config(self, wb: openpyxl.Workbook) -> Config:
        """Parse the optional config sheet (key-value pairs).

        Falls back to lite defaults if no config sheet is present.
        """
        ws = self._get_sheet(wb, "config", required=False)
        if not ws:
            return Config()

        rows = self._read_sheet_rows(ws)
        config_dict: dict[str, Any] = {}

        for row in rows:
            # Normalize header keys to lowercase for case-insensitive matching
            lc_row = {str(k).lower(): v for k, v in row.items()}
            key = _str(lc_row.get("key") or lc_row.get("config_key"))
            value = lc_row.get("value") or lc_row.get("config_value")
            if key and value is not None:
                config_dict[key.strip()] = value
            elif key:
                self.warnings.append(
                    f"Config key '{key}' has no value -- skipped"
                )

        if config_dict:
            self.warnings.append(
                f"Config sheet: read {len(config_dict)} settings: "
                + ", ".join(f"{k}={v}" for k, v in config_dict.items())
            )
        else:
            self.warnings.append(
                "Config sheet found but no valid key-value pairs read. "
                "Expected headers: 'key' and 'value'"
            )

        return _build_config_from_dict(config_dict)

    # ------------------------------------------------------------------
    # Low-level helpers
    # ------------------------------------------------------------------

    def _get_sheet(
        self, wb: openpyxl.Workbook, sheet_name: str, required: bool = True
    ) -> Optional[Worksheet]:
        """Get a worksheet by name (case-insensitive search)."""
        for name in wb.sheetnames:
            if name.lower() == sheet_name.lower():
                return wb[name]
        if required:
            self.errors.append(
                f"Required sheet '{sheet_name}' not found. "
                f"Available sheets: {wb.sheetnames}"
            )
        return None

    @staticmethod
    def _read_sheet_rows(ws: Worksheet) -> list[dict[str, Any]]:
        """Read all data rows from a worksheet as dicts keyed by header name."""
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).strip() if h else "" for h in rows[0]]
        result = []

        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue
            row_dict: dict[str, Any] = {}
            for col_idx, header in enumerate(headers):
                if header and col_idx < len(row):
                    row_dict[header] = row[col_idx]
            result.append(row_dict)

        return result


# --- Helper Functions ---


def _str(value: Any) -> str:
    """Convert to string, returning empty string for None."""
    if value is None:
        return ""
    return str(value).strip()




def _parse_cardinality(value: str) -> tuple[Cardinality, bool]:
    """Parse relationship type string to Cardinality enum.

    Returns (cardinality, defaulted) where defaulted is True if the value
    was not recognized and the default 1:N was used.
    """
    normalized = value.strip().upper().replace(" ", "")
    if normalized in ("1:1", "1-1", "ONE_TO_ONE"):
        return Cardinality.ONE_TO_ONE, False
    if normalized in ("1:N", "1-N", "1:MANY", "ONE_TO_MANY"):
        return Cardinality.ONE_TO_MANY, False
    if normalized in ("M:N", "M-N", "MANY_TO_MANY", "N:M"):
        return Cardinality.MANY_TO_MANY, False
    return Cardinality.ONE_TO_MANY, True


def _parse_join_attributes(join_str: str) -> tuple[list[str], list[str]]:
    """Parse JoinAttributes string into parent and child column lists.

    Format: "ParentEntity.AttrName = ChildEntity.AttrName"
    Single join only (no AND clause needed per requirements).

    Returns:
        Tuple of (parent_key_columns, child_key_columns) using the
        attribute Code (the part after the dot).
    """
    if not join_str or not join_str.strip():
        return [], []

    join_str = join_str.strip()

    # Split on '='
    parts = join_str.split("=")
    if len(parts) != 2:
        return [], []

    left = parts[0].strip()
    right = parts[1].strip()

    # Each side is "EntityName.Attribute Name"
    # Split on first '.' only (attribute names could contain dots, but unlikely)
    left_attr = _extract_attr_from_join_part(left)
    right_attr = _extract_attr_from_join_part(right)

    if not left_attr or not right_attr:
        return [], []

    return [left_attr], [right_attr]


def _extract_attr_from_join_part(part: str) -> str:
    """Extract attribute name from 'EntityName.AttributeName' format.

    The attribute name in JoinAttributes uses the display Name (e.g., "Account Uuid")
    which corresponds to the Code column (e.g., "ACCOUNT_UUID").

    We convert the display name to the Code format:
    - Convert to uppercase
    - Replace spaces with underscores
    """
    dot_pos = part.find(".")
    if dot_pos < 0:
        return ""
    # Validate: must have content on both sides of the dot
    entity_part = part[:dot_pos].strip()
    if not entity_part:
        return ""
    attr_display_name = part[dot_pos + 1:].strip()
    if not attr_display_name:
        return ""
    # Convert display name to code format: uppercase, spaces -> underscores
    return attr_display_name.upper().replace(" ", "_")


def _build_name_map(canonical_names: list[str]) -> dict[str, str]:
    """Build a lowercase -> canonical name lookup for cross-sheet matching.

    Given the entity names from the Entities sheet (the source of truth),
    creates a dict mapping lowercased names to their original form.
    This allows case-insensitive matching from Attributes and Relationships.
    """
    return {name.lower(): name for name in canonical_names}


