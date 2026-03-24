"""Excel parser: reads input workbook and produces a LogicalModel."""

import re
from pathlib import Path
from typing import Any, Callable, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from spark_pdm_generator.models.logical import (
    Attribute,
    Cardinality,
    Config,
    DataDistribution,
    DenormalizationMode,
    Entity,
    EntityType,
    GrowthRate,
    LogicalModel,
    Priority,
    QueryFrequency,
    QueryPattern,
    Relationship,
    RuleOverride,
    SkewIndicator,
    UpdateFrequency,
)
from spark_pdm_generator.parsers.column_mapper import (
    ColumnMapping,
    SheetMapping,
    create_default_mapping,
)
from spark_pdm_generator.parsers.utils import (
    build_config_from_dict as _build_config_from_dict,
    parse_enum as _parse_enum,
    parse_float as _parse_float,
    parse_int as _parse_int,
    parse_string_list as _parse_string_list,
)


class ParseError(Exception):
    """Raised when the input workbook has structural issues."""

    pass


class ExcelParser:
    """Reads an input Excel workbook and produces a LogicalModel."""

    def __init__(self, mapping: Optional[ColumnMapping] = None) -> None:
        self.mapping = mapping or create_default_mapping()
        self.errors: list[str] = []
        self.warnings: list[str] = []

    def parse(self, workbook_path: Path) -> LogicalModel:
        """Parse the input workbook into a LogicalModel.

        Args:
            workbook_path: Path to the Excel file.

        Returns:
            A populated LogicalModel.

        Raises:
            ParseError: If required sheets or fields are missing.
        """
        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

        try:
            entities = self._parse_entities(wb)
            # Build canonical name map for cross-sheet normalization
            self._entity_name_map: dict[str, str] = {
                e.entity_name.lower(): e.entity_name for e in entities
            }
            attributes = self._parse_attributes(wb)
            relationships = self._parse_relationships(wb)
            distributions = self._parse_data_distribution(wb)
            query_patterns = self._parse_query_patterns(wb)
            rule_overrides = self._parse_rules_override(wb)
            config = self._parse_config(wb)
        finally:
            wb.close()

        # Warn on duplicate entity names
        seen_entities: set[str] = set()
        for e in entities:
            if e.entity_name in seen_entities:
                self.warnings.append(
                    f"Duplicate entity name: '{e.entity_name}'"
                )
            seen_entities.add(e.entity_name)

        # Warn on duplicate (entity, attribute) pairs
        seen_attrs: set[tuple[str, str]] = set()
        for a in attributes:
            key = (a.entity_name, a.attribute_name)
            if key in seen_attrs:
                self.warnings.append(
                    f"Duplicate attribute: '{a.entity_name}.{a.attribute_name}'"
                )
            seen_attrs.add(key)

        if self.errors:
            raise ParseError(
                "Parsing errors:\n" + "\n".join(f"  - {e}" for e in self.errors)
            )

        return LogicalModel(
            entities=entities,
            attributes=attributes,
            relationships=relationships,
            data_distributions=distributions,
            query_patterns=query_patterns,
            rule_overrides=rule_overrides,
            config=config,
        )

    # ------------------------------------------------------------------
    # Generic sheet parser
    # ------------------------------------------------------------------

    def _parse_sheet(
        self,
        wb: openpyxl.Workbook,
        sheet_name: str,
        *,
        required: bool,
        use_mapping: bool,
        required_fields: tuple[str, ...],
        builder_fn: Callable[[dict[str, Any], int], Any],
        missing_warning: str = "",
    ) -> list:
        """Generic sheet-parsing skeleton shared by all _parse_* methods.

        Args:
            wb: The open workbook.
            sheet_name: Internal sheet name (may be mapped).
            required: If True, missing required fields produce errors;
                      if False, rows with missing fields are silently skipped.
            use_mapping: Whether to use column-name mapping for this sheet.
            required_fields: Field names that must be non-empty.
            builder_fn: (row_dict, row_number) -> model object or None.
            missing_warning: If non-empty and the sheet is absent, append
                             this as a parser warning.

        Returns:
            A list of model objects built by *builder_fn*.
        """
        ws = self._get_sheet(wb, sheet_name, required=required)
        if not ws:
            if missing_warning:
                self.warnings.append(missing_warning)
            return []

        mapping = self.mapping.get_sheet_mapping(sheet_name) if use_mapping else None
        rows = self._read_sheet_rows(ws, mapping)
        results = []

        for i, row in enumerate(rows, start=2):
            # Validate required fields
            missing = [f for f in required_fields if not row.get(f)]
            if missing:
                if required:
                    for f in missing:
                        self.errors.append(f"{sheet_name} row {i}: {f} is required")
                continue

            obj = builder_fn(row, i)
            if obj is not None:
                results.append(obj)

        return results

    # ------------------------------------------------------------------
    # Sheet-specific builders
    # ------------------------------------------------------------------

    def _parse_entities(self, wb: openpyxl.Workbook) -> list[Entity]:
        return self._parse_sheet(
            wb, "entities", required=True, use_mapping=True,
            required_fields=("entity_name",),
            builder_fn=self._build_entity,
        )

    @staticmethod
    def _build_entity(row: dict[str, Any], i: int) -> Entity:
        return Entity(
            entity_name=str(row["entity_name"]).strip(),
            entity_type=_parse_enum(
                row.get("entity_type"), EntityType, EntityType.UNKNOWN
            ),
            description=str(row.get("description", "") or ""),
            grain_description=str(row.get("grain_description", "") or ""),
            domain=str(row.get("domain", "general") or "general"),
            estimated_row_count=_parse_int(row.get("estimated_row_count")),
            estimated_record_length_bytes=_parse_int(
                row.get("estimated_record_length_bytes")
            ),
            growth_rate=_parse_enum(
                row.get("growth_rate"), GrowthRate, GrowthRate.MODERATE
            ),
            update_frequency=_parse_enum(
                row.get("update_frequency"),
                UpdateFrequency,
                UpdateFrequency.DAILY,
            ),
        )

    def _parse_attributes(self, wb: openpyxl.Workbook) -> list[Attribute]:
        attrs = self._parse_sheet(
            wb, "attributes", required=True, use_mapping=True,
            required_fields=("entity_name", "attribute_name", "logical_data_type"),
            builder_fn=self._build_attribute,
        )
        # Normalize entity names to match canonical names from Entities sheet
        for attr in attrs:
            canonical = self._entity_name_map.get(attr.entity_name.lower())
            if canonical and canonical != attr.entity_name:
                attr.entity_name = canonical
        return attrs

    @staticmethod
    def _build_attribute(row: dict[str, Any], i: int) -> Attribute:
        data_type = row["logical_data_type"]
        precision, scale = _parse_type_precision(str(data_type))
        row_precision = _parse_int(row.get("precision"))
        row_scale = _parse_int(row.get("scale"))

        return Attribute(
            entity_name=str(row["entity_name"]).strip(),
            attribute_name=str(row["attribute_name"]).strip(),
            logical_data_type=_normalize_data_type(str(data_type)),
            precision=row_precision or precision,
            scale=row_scale or scale,
            max_length=_parse_int(row.get("max_length")),
            nullable=_parse_bool(row.get("nullable"), default=True),
            is_primary_key=_parse_bool(row.get("is_primary_key"), default=False),
            is_foreign_key=_parse_bool(row.get("is_foreign_key"), default=False),
            fk_references=_parse_optional_str(row.get("fk_references")),
            description=str(row.get("description", "") or ""),
            domain_group=_parse_optional_str(row.get("domain_group")),
        )

    def _parse_relationships(self, wb: openpyxl.Workbook) -> list[Relationship]:
        rels = self._parse_sheet(
            wb, "relationships", required=True, use_mapping=True,
            required_fields=(
                "parent_entity", "child_entity", "cardinality",
                "parent_key_columns", "child_key_columns",
            ),
            builder_fn=self._build_relationship,
        )
        # Normalize entity names to match canonical names from Entities sheet
        for rel in rels:
            canonical_parent = self._entity_name_map.get(rel.parent_entity.lower())
            canonical_child = self._entity_name_map.get(rel.child_entity.lower())
            if canonical_parent and canonical_parent != rel.parent_entity:
                rel.parent_entity = canonical_parent
            if canonical_child and canonical_child != rel.child_entity:
                rel.child_entity = canonical_child
        return rels

    @staticmethod
    def _build_relationship(row: dict[str, Any], i: int) -> Relationship:
        return Relationship(
            parent_entity=str(row["parent_entity"]).strip(),
            child_entity=str(row["child_entity"]).strip(),
            cardinality=_parse_enum(
                row["cardinality"], Cardinality, Cardinality.ONE_TO_MANY
            ),
            parent_key_columns=_parse_string_list(str(row["parent_key_columns"])),
            child_key_columns=_parse_string_list(str(row["child_key_columns"])),
            is_identifying=_parse_bool(
                row.get("is_identifying"), default=False
            ),
            description=str(row.get("description", "") or ""),
        )

    def _parse_data_distribution(
        self, wb: openpyxl.Workbook
    ) -> list[DataDistribution]:
        dists = self._parse_sheet(
            wb, "data_distribution", required=False, use_mapping=False,
            required_fields=("entity_name", "attribute_name"),
            builder_fn=self._build_distribution,
            missing_warning=(
                "data_distribution sheet not found. "
                "Tool will use heuristics instead of data-driven decisions."
            ),
        )
        # Normalize entity names
        for d in dists:
            canonical = self._entity_name_map.get(d.entity_name.lower())
            if canonical and canonical != d.entity_name:
                d.entity_name = canonical
        return dists

    @staticmethod
    def _build_distribution(row: dict[str, Any], i: int) -> DataDistribution:
        return DataDistribution(
            entity_name=str(row["entity_name"]).strip(),
            attribute_name=str(row["attribute_name"]).strip(),
            distinct_count=_parse_int(row.get("distinct_count")),
            null_percentage=_parse_float(row.get("null_percentage"), 0.0),
            min_value=_parse_optional_str(row.get("min_value")),
            max_value=_parse_optional_str(row.get("max_value")),
            avg_length=_parse_float(row.get("avg_length"), None),
            skew_indicator=_parse_enum(
                row.get("skew_indicator"), SkewIndicator, SkewIndicator.LOW
            ),
        )

    def _parse_query_patterns(self, wb: openpyxl.Workbook) -> list[QueryPattern]:
        return self._parse_sheet(
            wb, "query_patterns", required=False, use_mapping=False,
            required_fields=("pattern_name", "primary_entity"),
            builder_fn=self._build_query_pattern,
            missing_warning=(
                "query_patterns sheet not found. "
                "Tool will produce generic optimization only."
            ),
        )

    @staticmethod
    def _build_query_pattern(row: dict[str, Any], i: int) -> QueryPattern:
        return QueryPattern(
            pattern_name=str(row["pattern_name"]).strip(),
            primary_entity=str(row["primary_entity"]).strip(),
            filter_attributes=_parse_string_list(
                str(row.get("filter_attributes", "") or "")
            ),
            group_by_attributes=_parse_string_list(
                str(row.get("group_by_attributes", "") or "")
            ),
            join_entities=_parse_string_list(
                str(row.get("join_entities", "") or "")
            ),
            accessed_attributes=_parse_string_list(
                str(row.get("accessed_attributes", "") or "")
            ),
            frequency=_parse_enum(
                row.get("frequency"), QueryFrequency, QueryFrequency.DAILY
            ),
            priority=_parse_enum(
                row.get("priority"), Priority, Priority.MEDIUM
            ),
        )

    def _parse_rules_override(self, wb: openpyxl.Workbook) -> list[RuleOverride]:
        return self._parse_sheet(
            wb, "rules_override", required=False, use_mapping=False,
            required_fields=("override_type", "target", "instruction"),
            builder_fn=self._build_rule_override,
        )

    @staticmethod
    def _build_rule_override(row: dict[str, Any], i: int) -> Optional[RuleOverride]:
        rule_id = _parse_int(row.get("rule_id"))
        if rule_id is None:
            return None
        return RuleOverride(
            rule_id=rule_id,
            override_type=str(row["override_type"]).strip(),
            target=str(row["target"]).strip(),
            instruction=str(row["instruction"]).strip(),
        )

    # ------------------------------------------------------------------
    # Config (unique shape -- not genericized)
    # ------------------------------------------------------------------

    def _parse_config(self, wb: openpyxl.Workbook) -> Config:
        """Parse the config sheet (key-value pairs)."""
        ws = self._get_sheet(wb, "config", required=False)
        if not ws:
            return Config()

        rows = self._read_sheet_rows(ws)
        config_dict: dict[str, Any] = {}

        for row in rows:
            key = row.get("key") or row.get("config_key")
            value = row.get("value") or row.get("config_value")
            if key and value is not None:
                config_dict[str(key).strip()] = value

        return _build_config_from_dict(config_dict)

    # ------------------------------------------------------------------
    # Low-level helpers
    # ------------------------------------------------------------------

    def _get_sheet(
        self, wb: openpyxl.Workbook, internal_name: str, required: bool = True
    ) -> Optional[Worksheet]:
        """Get a worksheet by its mapped name."""
        tab_name = self.mapping.get_sheet_name(internal_name)
        if tab_name in wb.sheetnames:
            return wb[tab_name]
        if internal_name in wb.sheetnames:
            return wb[internal_name]
        if required:
            self.errors.append(
                f"Required sheet '{tab_name}' (mapped from '{internal_name}') not found. "
                f"Available sheets: {wb.sheetnames}"
            )
        return None

    def _read_sheet_rows(
        self,
        ws: Worksheet,
        sheet_mapping: Optional[SheetMapping] = None,
    ) -> list[dict[str, Any]]:
        """Read all data rows from a worksheet, mapping column headers.

        Returns a list of dicts keyed by internal field names.
        """
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        raw_headers = [str(h).strip() if h else "" for h in rows[0]]

        if sheet_mapping:
            header_to_internal: dict[int, str] = {}
            for col_idx, header in enumerate(raw_headers):
                if not header:
                    continue
                internal = sheet_mapping.get_internal_field(header)
                if internal:
                    header_to_internal[col_idx] = internal
        else:
            header_to_internal = {i: h for i, h in enumerate(raw_headers) if h}

        result = []
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue
            row_dict: dict[str, Any] = {}
            for col_idx, internal_field in header_to_internal.items():
                if col_idx < len(row):
                    row_dict[internal_field] = row[col_idx]
            result.append(row_dict)

        return result


# --- Helper Functions ---




def _parse_bool(value: Any, default: bool) -> bool:
    """Parse a value to bool. Accepts Y/N, True/False, 1/0."""
    if value is None:
        return default
    str_val = str(value).strip().upper()
    if str_val in ("Y", "YES", "TRUE", "1"):
        return True
    if str_val in ("N", "NO", "FALSE", "0"):
        return False
    return default


def _parse_optional_str(value: Any) -> Optional[str]:
    """Parse to string or None."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None




def _parse_type_precision(data_type: str) -> tuple[Optional[int], Optional[int]]:
    """Extract precision and scale from a type string like 'DECIMAL(15,2)'.

    Handles whitespace variants: 'DECIMAL( 15 , 2 )', 'DECIMAL(15,2)', etc.
    Returns (precision, scale) or (None, None).
    """
    match = re.match(r".*\(\s*(\d+)\s*,\s*(\d+)\s*\)", data_type)
    if match:
        return int(match.group(1)), int(match.group(2))
    match = re.match(r".*\(\s*(\d+)\s*\)", data_type)
    if match:
        return int(match.group(1)), None
    return None, None


def _normalize_data_type(data_type: str) -> str:
    """Normalize a data type string by removing precision/scale info.

    'DECIMAL(15,2)' -> 'DECIMAL'
    'VARCHAR(255)' -> 'VARCHAR'
    'INTEGER' -> 'INTEGER'
    """
    return re.sub(r"\(.*\)", "", data_type).strip().upper()


