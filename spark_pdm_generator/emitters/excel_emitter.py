"""Excel emitter: generates the output workbook with all 7 sheets."""

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from spark_pdm_generator.models.physical import PhysicalModel


# Header styling
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")


def emit_excel(output: PhysicalModel, output_path: Path) -> None:
    """Write the complete output workbook."""
    from spark_pdm_generator.emitters.ddl_builder import DDLBuilder

    wb = Workbook()
    wb.remove(wb.active)

    # --- physical_entities ---
    _write_sheet(wb, "physical_entities", [
        "physical_entity_name", "source_entities", "entity_type",
        "grain_description", "domain", "partition_columns",
        "bucket_column", "bucket_count", "sort_columns",
        "estimated_row_count", "estimated_size_mb",
        "estimated_file_count", "row_group_size_mb",
        "compression_codec", "storage_format", "denormalization_notes",
    ], [
        [
            e.physical_entity_name,
            ", ".join(e.source_entities),
            e.entity_type.value,
            e.grain_description,
            e.domain,
            ", ".join(e.partition_columns),
            e.bucket_column or "",
            e.bucket_count,
            ", ".join(f"{sc.column_name} {sc.order.value}" for sc in e.sort_columns),
            e.estimated_row_count,
            round(e.estimated_size_bytes / (1024 * 1024)) if e.estimated_size_bytes else None,
            e.estimated_file_count,
            e.row_group_size_mb,
            e.compression_codec,
            e.storage_format,
            e.denormalization_notes,
        ]
        for e in output.physical_entities
    ])

    # --- physical_attributes ---
    _write_sheet(wb, "physical_attributes", [
        "physical_entity_name", "attribute_name", "source_entity",
        "source_attribute", "parquet_type", "logical_type", "encoding",
        "nullable", "is_primary_key", "is_partition_col", "is_bucket_col",
        "is_sort_col", "sort_order", "estimated_cardinality",
        "estimated_null_pct", "notes",
    ], [
        [
            a.physical_entity_name, a.attribute_name, a.source_entity,
            a.source_attribute, a.parquet_type, a.logical_type,
            a.encoding.value if hasattr(a.encoding, "value") else str(a.encoding),
            "Y" if a.nullable else "N",
            "Y" if a.is_primary_key else "N",
            "Y" if a.is_partition_col else "N",
            "Y" if a.is_bucket_col else "N",
            "Y" if a.is_sort_col else "N",
            a.sort_order.value if a.sort_order else "",
            a.estimated_cardinality, a.estimated_null_pct, a.notes,
        ]
        for a in output.physical_attributes
    ])

    # --- physical_relationships ---
    _write_sheet(wb, "physical_relationships", [
        "parent_physical_entity", "child_physical_entity", "join_type",
        "join_columns", "co_partitioned", "co_bucketed",
        "estimated_join_cost", "notes",
    ], [
        [
            r.parent_physical_entity, r.child_physical_entity,
            r.join_type.value, ", ".join(r.join_columns),
            "Y" if r.co_partitioned else "N",
            "Y" if r.co_bucketed else "N",
            r.estimated_join_cost.value, r.notes,
        ]
        for r in output.physical_relationships
    ])

    # --- transformation_log ---
    _write_sheet(wb, "transformation_log", [
        "log_id", "rule_applied", "source_entity", "target_entity",
        "description", "rationale",
    ], [
        [
            t.log_id, t.rule_applied.value, t.source_entity,
            t.target_entity, t.description, t.rationale,
        ]
        for t in output.transformation_log
    ])

    # --- spark_ddl ---
    builder = DDLBuilder()
    _write_sheet(wb, "spark_ddl", [
        "physical_entity_name", "format", "ddl_statement",
    ], [
        [
            e.physical_entity_name, "PARQUET",
            builder.build_create_table_parquet(
                e, output.get_attributes_for_entity(e.physical_entity_name)
            ),
        ]
        for e in output.physical_entities
    ])

    # --- spark_config ---
    _write_sheet(wb, "spark_config", [
        "config_key", "config_value", "description",
    ], [
        [c.config_key, c.config_value, c.description]
        for c in output.spark_config
    ])

    # --- warnings ---
    _write_sheet(wb, "warnings", [
        "level", "entity", "attribute", "message", "recommendation",
    ], [
        [
            w.level.value, w.entity, w.attribute,
            w.message, w.recommendation,
        ]
        for w in output.warnings
    ])

    wb.save(output_path)


def _write_sheet(
    wb: Workbook,
    sheet_name: str,
    headers: list[str],
    rows: Iterable[list],
) -> None:
    """Create a sheet, write styled headers and data rows, auto-size columns."""
    ws = wb.create_sheet(sheet_name)
    _write_header(ws, headers)
    for row in rows:
        ws.append(row)
    _auto_width(ws)


def _write_header(ws, headers: list[str]) -> None:
    """Write styled header row."""
    ws.append(headers)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws, min_width: int = 10, max_width: int = 60) -> None:
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        lengths = []
        for cell in col:
            if cell.value:
                lengths.append(len(str(cell.value)))
        if lengths:
            width = min(max(max(lengths), min_width), max_width)
            ws.column_dimensions[col_letter].width = width + 2
