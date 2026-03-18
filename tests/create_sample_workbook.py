"""Creates a sample credit card issuer input workbook for testing."""

from pathlib import Path

from openpyxl import Workbook


def create_sample_workbook(output_path: Path) -> None:
    """Create a sample input workbook with a small credit card model."""
    wb = Workbook()
    wb.remove(wb.active)

    # --- entities sheet ---
    ws = wb.create_sheet("entities")
    ws.append([
        "entity_name", "entity_type", "description", "grain_description",
        "domain", "estimated_row_count", "estimated_record_length_bytes",
        "growth_rate", "update_frequency",
    ])
    entities = [
        ["card_account", "FACT", "Core account-level reporting entity", "One row per account per report_date", "account", 50_000_000, 500, "HIGH", "DAILY"],
        ["card_product", "DIMENSION", "Card product catalog", "One row per product", "reference", 47, 200, "STATIC", "STATIC"],
        ["merchant", "DIMENSION", "Merchant master data", "One row per merchant", "transaction", 2_000_000, 300, "MODERATE", "DAILY"],
        ["account_transaction_agg", "FACT", "Daily transaction aggregates per account", "One row per account per report_date", "transaction", 50_000_000, 200, "HIGH", "DAILY"],
        ["risk_score", "DIMENSION", "Risk scoring attributes", "One row per account per scoring_date", "risk", 50_000_000, 150, "HIGH", "DAILY"],
        ["balance_detail", "FACT", "Balance and financial detail", "One row per account per report_date", "financial", 50_000_000, 250, "HIGH", "DAILY"],
        ["collections_status", "DIMENSION", "Collections tracking", "One row per delinquent account per report_date", "collections", 2_500_000, 200, "MODERATE", "DAILY"],
        ["reward_program", "DIMENSION", "Rewards program reference", "One row per program", "reference", 12, 150, "STATIC", "STATIC"],
    ]
    for e in entities:
        ws.append(e)

    # --- attributes sheet ---
    ws = wb.create_sheet("attributes")
    ws.append([
        "entity_name", "attribute_name", "logical_data_type", "precision", "scale",
        "max_length", "nullable", "is_primary_key", "is_foreign_key",
        "fk_references", "description", "domain_group",
    ])
    attributes = [
        # card_account
        ["card_account", "account_id", "BIGINT", None, None, None, "N", "Y", "N", None, "Unique account identifier", "account"],
        ["card_account", "report_date", "DATE", None, None, None, "N", "Y", "N", None, "Reporting date", "account"],
        ["card_account", "product_id", "INTEGER", None, None, None, "N", "N", "Y", "card_product.product_id", "FK to product", "account"],
        ["card_account", "customer_name", "VARCHAR", None, None, 100, "Y", "N", "N", None, "Customer name", "account"],
        ["card_account", "credit_limit", "DECIMAL", 15, 2, None, "Y", "N", "N", None, "Credit limit amount", "financial"],
        ["card_account", "open_date", "DATE", None, None, None, "Y", "N", "N", None, "Account open date", "account"],
        ["card_account", "account_status", "VARCHAR", None, None, 20, "N", "N", "N", None, "Current account status", "account"],
        ["card_account", "state_code", "VARCHAR", None, None, 2, "Y", "N", "N", None, "Customer state", "account"],
        ["card_account", "reward_program_id", "INTEGER", None, None, None, "Y", "N", "Y", "reward_program.program_id", "FK to rewards", "account"],
        # card_product
        ["card_product", "product_id", "INTEGER", None, None, None, "N", "Y", "N", None, "Product identifier", None],
        ["card_product", "product_name", "VARCHAR", None, None, 100, "N", "N", "N", None, "Product name", None],
        ["card_product", "product_type", "VARCHAR", None, None, 30, "N", "N", "N", None, "Product type (VISA, MC, etc.)", None],
        ["card_product", "annual_fee", "DECIMAL", 10, 2, None, "Y", "N", "N", None, "Annual fee amount", None],
        # merchant
        ["merchant", "merchant_id", "BIGINT", None, None, None, "N", "Y", "N", None, "Merchant identifier", None],
        ["merchant", "merchant_name", "VARCHAR", None, None, 200, "Y", "N", "N", None, "Merchant name", None],
        ["merchant", "mcc_code", "VARCHAR", None, None, 4, "N", "N", "N", None, "Merchant category code", None],
        ["merchant", "merchant_state", "VARCHAR", None, None, 2, "Y", "N", "N", None, "Merchant state", None],
        # account_transaction_agg
        ["account_transaction_agg", "account_id", "BIGINT", None, None, None, "N", "Y", "N", None, "Account identifier", "transaction"],
        ["account_transaction_agg", "report_date", "DATE", None, None, None, "N", "Y", "N", None, "Reporting date", "transaction"],
        ["account_transaction_agg", "total_spend", "DECIMAL", 15, 2, None, "Y", "N", "N", None, "Total daily spend", "transaction"],
        ["account_transaction_agg", "transaction_count", "INTEGER", None, None, None, "Y", "N", "N", None, "Number of transactions", "transaction"],
        ["account_transaction_agg", "avg_ticket_size", "DECIMAL", 10, 2, None, "Y", "N", "N", None, "Average transaction amount", "transaction"],
        ["account_transaction_agg", "top_mcc_code", "VARCHAR", None, None, 4, "Y", "N", "N", None, "Most frequent MCC", "transaction"],
        # risk_score
        ["risk_score", "account_id", "BIGINT", None, None, None, "N", "Y", "N", None, "Account identifier", "risk"],
        ["risk_score", "scoring_date", "DATE", None, None, None, "N", "Y", "N", None, "Score date", "risk"],
        ["risk_score", "fico_score", "INTEGER", None, None, None, "Y", "N", "N", None, "FICO score", "risk"],
        ["risk_score", "behavioral_score", "DECIMAL", 5, 2, None, "Y", "N", "N", None, "Internal behavioral score", "risk"],
        ["risk_score", "delinquency_bucket", "VARCHAR", None, None, 10, "Y", "N", "N", None, "Delinquency bucket (CURRENT, 30, 60, 90, 120+)", "risk"],
        ["risk_score", "probability_of_default", "DECIMAL", 7, 6, None, "Y", "N", "N", None, "PD model output", "risk"],
        # balance_detail
        ["balance_detail", "account_id", "BIGINT", None, None, None, "N", "Y", "N", None, "Account identifier", "financial"],
        ["balance_detail", "report_date", "DATE", None, None, None, "N", "Y", "N", None, "Reporting date", "financial"],
        ["balance_detail", "current_balance", "DECIMAL", 15, 2, None, "Y", "N", "N", None, "Current balance", "financial"],
        ["balance_detail", "statement_balance", "DECIMAL", 15, 2, None, "Y", "N", "N", None, "Statement balance", "financial"],
        ["balance_detail", "minimum_payment_due", "DECIMAL", 10, 2, None, "Y", "N", "N", None, "Minimum payment due", "financial"],
        ["balance_detail", "interest_accrued", "DECIMAL", 10, 2, None, "Y", "N", "N", None, "Interest accrued", "financial"],
        ["balance_detail", "payment_amount", "DECIMAL", 15, 2, None, "Y", "N", "N", None, "Last payment amount", "financial"],
        # collections_status
        ["collections_status", "account_id", "BIGINT", None, None, None, "N", "Y", "N", None, "Account identifier", "collections"],
        ["collections_status", "report_date", "DATE", None, None, None, "N", "Y", "N", None, "Reporting date", "collections"],
        ["collections_status", "collection_status", "VARCHAR", None, None, 30, "Y", "N", "N", None, "Collection status", "collections"],
        ["collections_status", "promise_to_pay", "BOOLEAN", None, None, None, "Y", "N", "N", None, "PTP flag", "collections"],
        ["collections_status", "recovery_amount", "DECIMAL", 15, 2, None, "Y", "N", "N", None, "Recovery amount", "collections"],
        # reward_program
        ["reward_program", "program_id", "INTEGER", None, None, None, "N", "Y", "N", None, "Program identifier", None],
        ["reward_program", "program_name", "VARCHAR", None, None, 100, "N", "N", "N", None, "Program name", None],
        ["reward_program", "points_multiplier", "DECIMAL", 5, 2, None, "Y", "N", "N", None, "Points multiplier", None],
    ]
    for a in attributes:
        ws.append(a)

    # --- relationships sheet ---
    ws = wb.create_sheet("relationships")
    ws.append([
        "parent_entity", "child_entity", "cardinality",
        "parent_key_columns", "child_key_columns", "is_identifying", "description",
    ])
    relationships = [
        ["card_product", "card_account", "1:N", "product_id", "product_id", "N", "Product to account"],
        ["reward_program", "card_account", "1:N", "program_id", "reward_program_id", "N", "Reward program to account"],
        ["card_account", "account_transaction_agg", "1:N", "account_id, report_date", "account_id, report_date", "Y", "Account to daily txn agg"],
        ["card_account", "risk_score", "1:N", "account_id", "account_id", "N", "Account to risk scores"],
        ["card_account", "balance_detail", "1:N", "account_id, report_date", "account_id, report_date", "Y", "Account to balance detail"],
        ["card_account", "collections_status", "1:N", "account_id, report_date", "account_id, report_date", "N", "Account to collections"],
    ]
    for r in relationships:
        ws.append(r)

    # --- data_distribution sheet ---
    ws = wb.create_sheet("data_distribution")
    ws.append([
        "entity_name", "attribute_name", "distinct_count",
        "null_percentage", "min_value", "max_value", "avg_length", "skew_indicator",
    ])
    distributions = [
        ["card_account", "account_id", 50_000_000, 0.0, "1", "50000000", 8, "LOW"],
        ["card_account", "report_date", 365, 0.0, "2024-01-01", "2024-12-31", 10, "LOW"],
        ["card_account", "product_id", 47, 0.0, "1", "47", 2, "MODERATE"],
        ["card_account", "account_status", 5, 0.0, "ACTIVE", "CLOSED", 8, "HIGH"],
        ["card_account", "state_code", 52, 2.0, "AK", "WY", 2, "MODERATE"],
        ["card_account", "credit_limit", 5000, 0.5, "500.00", "100000.00", 8, "MODERATE"],
        ["card_product", "product_id", 47, 0.0, "1", "47", 2, "LOW"],
        ["card_product", "product_type", 4, 0.0, "VISA", "AMEX", 6, "MODERATE"],
        ["merchant", "merchant_id", 2_000_000, 0.0, "1", "2000000", 8, "LOW"],
        ["merchant", "mcc_code", 400, 0.0, "0001", "9999", 4, "MODERATE"],
        ["account_transaction_agg", "account_id", 50_000_000, 0.0, "1", "50000000", 8, "LOW"],
        ["account_transaction_agg", "report_date", 365, 0.0, "2024-01-01", "2024-12-31", 10, "LOW"],
        ["account_transaction_agg", "top_mcc_code", 400, 5.0, "0001", "9999", 4, "MODERATE"],
        ["risk_score", "account_id", 50_000_000, 0.0, "1", "50000000", 8, "LOW"],
        ["risk_score", "delinquency_bucket", 5, 0.0, "CURRENT", "120+", 7, "HIGH"],
        ["balance_detail", "account_id", 50_000_000, 0.0, "1", "50000000", 8, "LOW"],
        ["balance_detail", "report_date", 365, 0.0, "2024-01-01", "2024-12-31", 10, "LOW"],
        ["collections_status", "account_id", 2_500_000, 0.0, "1", "50000000", 8, "MODERATE"],
        ["collections_status", "collection_status", 8, 0.0, "NEW", "CLOSED", 10, "MODERATE"],
    ]
    for d in distributions:
        ws.append(d)

    # --- query_patterns sheet ---
    ws = wb.create_sheet("query_patterns")
    ws.append([
        "pattern_name", "primary_entity", "filter_attributes",
        "group_by_attributes", "join_entities", "accessed_attributes",
        "frequency", "priority",
    ])
    patterns = [
        ["daily_account_summary", "card_account", "report_date", "product_id, state_code", "balance_detail, risk_score", "account_id, credit_limit, account_status, current_balance, fico_score", "DAILY", "HIGH"],
        ["delinquency_report", "card_account", "report_date, delinquency_bucket", "product_id, delinquency_bucket", "risk_score, balance_detail, collections_status", "account_id, fico_score, current_balance, collection_status", "DAILY", "HIGH"],
        ["transaction_analysis", "account_transaction_agg", "report_date", "top_mcc_code", "", "account_id, total_spend, transaction_count, avg_ticket_size", "WEEKLY", "MEDIUM"],
        ["full_regulatory_extract", "card_account", "report_date", "", "account_transaction_agg, risk_score, balance_detail, collections_status", "all", "MONTHLY", "HIGH"],
    ]
    for p in patterns:
        ws.append(p)

    # --- config sheet ---
    ws = wb.create_sheet("config")
    ws.append(["key", "value"])
    ws.append(["model_type", "OLTP_3NF"])
    ws.append(["target_format", "BOTH"])
    ws.append(["compression", "ZSTD"])
    ws.append(["cluster_parallelism", 2048])
    ws.append(["target_file_size_mb", 512])
    ws.append(["small_dim_row_threshold", 100000])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Sample workbook written to: {output_path}")


def create_aggressive_sample_workbook(output_path: Path) -> None:
    """Create a sample workbook configured for aggressive denormalization.

    Simulates a domain with 15 collections flattened into one wide table.
    """
    wb = Workbook()
    wb.remove(wb.active)

    # --- entities sheet ---
    ws = wb.create_sheet("entities")
    ws.append([
        "entity_name", "entity_type", "description", "grain_description",
        "domain", "estimated_row_count", "estimated_record_length_bytes",
        "growth_rate", "update_frequency",
    ])
    entities = [
        ["card_account", "FACT", "Core account-level reporting entity", "One row per account per report_date", "account", 300_000_000, 32_000, "HIGH", "DAILY"],
        ["card_product", "DIMENSION", "Card product catalog", "One row per product", "reference", 47, 200, "STATIC", "STATIC"],
        ["merchant", "DIMENSION", "Merchant master data", "One row per merchant", "transaction", 2_000_000, 300, "MODERATE", "DAILY"],
        ["risk_score", "DIMENSION", "Risk scoring attributes", "One row per account", "risk", 50_000_000, 150, "HIGH", "DAILY"],
        ["reward_program", "DIMENSION", "Rewards program reference", "One row per program", "reference", 12, 150, "STATIC", "STATIC"],
    ]
    for e in entities:
        ws.append(e)

    # --- attributes sheet ---
    ws = wb.create_sheet("attributes")
    ws.append([
        "entity_name", "attribute_name", "logical_data_type", "precision", "scale",
        "max_length", "nullable", "is_primary_key", "is_foreign_key",
        "fk_references", "description", "domain_group",
    ])
    attributes = [
        # card_account (central fact)
        ["card_account", "account_id", "BIGINT", None, None, None, "N", "Y", "N", None, "Unique account identifier", "account"],
        ["card_account", "report_date", "DATE", None, None, None, "N", "Y", "N", None, "Reporting date", "account"],
        ["card_account", "product_id", "INTEGER", None, None, None, "N", "N", "Y", "card_product.product_id", "FK to product", "account"],
        ["card_account", "customer_name", "VARCHAR", None, None, 100, "Y", "N", "N", None, "Customer name", "account"],
        ["card_account", "credit_limit", "DECIMAL", 15, 2, None, "Y", "N", "N", None, "Credit limit amount", "financial"],
        ["card_account", "reward_program_id", "INTEGER", None, None, None, "Y", "N", "Y", "reward_program.program_id", "FK to rewards", "account"],
        # card_product (small dim)
        ["card_product", "product_id", "INTEGER", None, None, None, "N", "Y", "N", None, "Product identifier", None],
        ["card_product", "product_name", "VARCHAR", None, None, 100, "N", "N", "N", None, "Product name", None],
        ["card_product", "product_type", "VARCHAR", None, None, 30, "N", "N", "N", None, "Product type", None],
        ["card_product", "annual_fee", "DECIMAL", 10, 2, None, "Y", "N", "N", None, "Annual fee", None],
        # merchant (large dim -- will be absorbed in aggressive mode)
        ["merchant", "merchant_id", "BIGINT", None, None, None, "N", "Y", "N", None, "Merchant identifier", None],
        ["merchant", "merchant_name", "VARCHAR", None, None, 200, "Y", "N", "N", None, "Merchant name", None],
        ["merchant", "mcc_code", "VARCHAR", None, None, 4, "N", "N", "N", None, "MCC code", None],
        # risk_score (large dim -- will be absorbed in aggressive mode)
        ["risk_score", "account_id", "BIGINT", None, None, None, "N", "Y", "N", None, "Account identifier", "risk"],
        ["risk_score", "fico_score", "INTEGER", None, None, None, "Y", "N", "N", None, "FICO score", "risk"],
        ["risk_score", "behavioral_score", "DECIMAL", 5, 2, None, "Y", "N", "N", None, "Behavioral score", "risk"],
        # reward_program (small dim)
        ["reward_program", "program_id", "INTEGER", None, None, None, "N", "Y", "N", None, "Program identifier", None],
        ["reward_program", "program_name", "VARCHAR", None, None, 100, "N", "N", "N", None, "Program name", None],
        ["reward_program", "points_multiplier", "DECIMAL", 5, 2, None, "Y", "N", "N", None, "Points multiplier", None],
    ]
    for a in attributes:
        ws.append(a)

    # --- relationships sheet ---
    ws = wb.create_sheet("relationships")
    ws.append([
        "parent_entity", "child_entity", "cardinality",
        "parent_key_columns", "child_key_columns", "is_identifying", "description",
    ])
    relationships = [
        ["card_product", "card_account", "1:N", "product_id", "product_id", "N", "Product to account"],
        ["reward_program", "card_account", "1:N", "program_id", "reward_program_id", "N", "Reward program to account"],
        ["merchant", "card_account", "1:N", "merchant_id", "account_id", "N", "Merchant to account"],
        ["risk_score", "card_account", "1:1", "account_id", "account_id", "N", "Risk score to account"],
    ]
    for r in relationships:
        ws.append(r)

    # --- data_distribution sheet ---
    ws = wb.create_sheet("data_distribution")
    ws.append([
        "entity_name", "attribute_name", "distinct_count",
        "null_percentage", "min_value", "max_value", "avg_length", "skew_indicator",
    ])
    distributions = [
        ["card_account", "account_id", 300_000_000, 0.0, "1", "300000000", 8, "LOW"],
        ["card_account", "report_date", 1, 0.0, "2025-01-15", "2025-01-15", 10, "LOW"],
        ["card_account", "product_id", 47, 0.0, "1", "47", 2, "MODERATE"],
        ["card_product", "product_id", 47, 0.0, "1", "47", 2, "LOW"],
        ["merchant", "merchant_id", 2_000_000, 0.0, "1", "2000000", 8, "LOW"],
    ]
    for d in distributions:
        ws.append(d)

    # --- query_patterns sheet ---
    ws = wb.create_sheet("query_patterns")
    ws.append([
        "pattern_name", "primary_entity", "filter_attributes",
        "group_by_attributes", "join_entities", "accessed_attributes",
        "frequency", "priority",
    ])
    patterns = [
        ["daily_report", "card_account", "report_date", "product_id", "", "all", "DAILY", "HIGH"],
    ]
    for p in patterns:
        ws.append(p)

    # --- config sheet (aggressive mode) ---
    ws = wb.create_sheet("config")
    ws.append(["key", "value"])
    ws.append(["model_type", "OLTP_3NF"])
    ws.append(["target_format", "BOTH"])
    ws.append(["compression", "ZSTD"])
    ws.append(["denormalization_mode", "AGGRESSIVE"])
    ws.append(["cluster_parallelism", 2048])
    ws.append(["target_file_size_mb", 2048])
    ws.append(["column_threshold_for_vertical_split", 3000])
    ws.append(["small_dim_row_threshold", 100000])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Aggressive-mode sample workbook written to: {output_path}")


if __name__ == "__main__":
    create_sample_workbook(Path("tests/sample_input.xlsx"))
    create_aggressive_sample_workbook(Path("tests/sample_aggressive_input.xlsx"))
